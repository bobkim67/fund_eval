"""evaluation_kis xlsx + pension_universe_kis 데이터를 snapshot.json 으로 변환.

산출:
  api/snapshot.json — frontend 가 fetch 해서 client-side 점수 재계산
구조:
  {
    "generated_at": "...",
    "filter_summary": {...},
    "sector_groups": ["국내주식", "국내채권", ...],
    "funds": [
      {
        "fund_cd": "K55101DX4743",
        "amc_nm": "한국투자신탁운용",
        "fund_nm": "...",
        "kis_type": "해외_주식혼합형",
        "sector_group": "혼합형",
        "aum": 195611908632,
        "yield_2y": 84.91,
        "sharp_2y": 1.76,
        "fam_aek": 425146833477,
        "amc_sector_aum_y": 123456789,
        "amc_sector_score_y": 96.5,
        "is_online": true,
        "is_pension": true,
        "haeji_normal": true,
        "filter_pass": true
      }, ...
    ]
  }
"""
import json
import datetime
import openpyxl
from pathlib import Path

DIR = Path(__file__).resolve().parent.parent.parent  # 펀드 리스트
SRC = DIR / "한투운용 퇴직연금펀드 리스트_20260518_FINAL.xlsx"
OUT = Path(__file__).resolve().parent / "snapshot.json"

# Sector 매핑 (Strict 4 + 혼합 + 기타)
def sector_group(kis_type: str) -> str:
    if not kis_type:
        return "기타"
    s = kis_type
    if s == "국내_주식형": return "국내주식"
    if s == "국내_채권형": return "국내채권"
    if s == "해외_주식형": return "해외주식"
    if s == "해외_채권형": return "해외채권"
    if "혼합" in s: return "혼합형"
    return "기타"


print(f"Loading {SRC.name} ...")
wb = openpyxl.load_workbook(SRC, data_only=True)
us = wb["pension_universe_kis"]
hdr = [us.cell(1, c).value for c in range(1, us.max_column + 1)]

def col(name):
    return hdr.index(name) + 1

# 필수 컬럼
COLS = {
    "fund_cd": "FUND_CD",
    "amc_nm": "AMC_NM",
    "fund_nm": "FUND_FNM_DB",
    "fund_nm_kis": "FUND_NM_KIS",
    "kis_type": "TYPE",
    "in_kis_lineup": "IN_KIS_LINEUP",
    "haeji": "HAEJI_GB",
    "class_gb": "CLASS_GB",
    "tmnt_gb": "TMNT_GB",
    "aum": "AUM",
    "nav": "NAV",
    "fam_aek": "FAM_AEK",
    "fam_size": "FAM_SIZE",
    "yield_2y": "YIELD_2Y",
    "yield_3y": "YIELD_3Y",
    "sharp_2y": "SHARP_2Y",
    "sharp_3y": "SHARP_3Y",
    "ir_2y": "IR_2Y",
    "te_2y": "TE_2Y",
    "alpha_3y": "ALPHA_3Y",
    "beta_3y": "BETA_3Y",
    "amc_sector_aum_y": "AMC_SECTOR_AUM_Y",
    "amc_sector_score_y": "AMC_SECTOR_SCORE_Y",
    "is_online": "IS_ONLINE",
    "is_pension_attb": "IS_PENSION_ATTB",
    "filter_pass": "FILTER_PASS",
    "fee_amc": "FEE_AMC",
    "fee_distb": "FEE_DISTB",
    "fee_total": "FEE_TOTAL",
}
col_idx = {k: col(v) for k, v in COLS.items()}

# 데이터 추출 — 서버 단 pre-filter: 정상 + 온라인 + 퇴직연금 만 포함
funds = []
sector_set = set()
total_rows = 0
skipped = 0
for r in range(3, us.max_row + 1):
    total_rows += 1
    row = {k: us.cell(r, ci).value for k, ci in col_idx.items()}
    sg = sector_group(row.get("kis_type"))
    is_online = row.get("is_online") == "Y"
    is_pension = row.get("is_pension_attb") == "Y"
    haeji_normal = row.get("haeji") == "1"
    # ★ 서버 단 pre-filter
    if not (haeji_normal and is_online and is_pension):
        skipped += 1
        continue
    sector_set.add(sg)
    filter_pass = row.get("filter_pass") == "Y"

    def f(v):
        if v is None: return None
        try: return float(v)
        except (TypeError, ValueError): return None

    # TDF 빈티지 추출 (펀드명에서 2015/2020/.../2060 등)
    import re as _re
    fnm = row.get("fund_nm") or row.get("fund_nm_kis") or ""
    has_tdf = "TDF" in fnm.upper()
    vintage_match = _re.search(r"(20\d{2})", fnm) if has_tdf else None
    tdf_vintage = vintage_match.group(1) if vintage_match else None

    funds.append({
        "fund_cd": row["fund_cd"],
        "is_tdf": has_tdf,
        "tdf_vintage": tdf_vintage,
        "amc_nm": row["amc_nm"],
        "fund_nm": row.get("fund_nm") or row.get("fund_nm_kis"),
        "kis_type": row.get("kis_type"),
        "sector_group": sg,
        "in_kis_lineup": row.get("in_kis_lineup"),
        "class_gb": row.get("class_gb"),
        "tmnt_gb": row.get("tmnt_gb"),
        "haeji_normal": haeji_normal,
        "is_online": is_online,
        "is_pension": is_pension,
        "filter_pass": filter_pass,
        "aum": f(row.get("aum")),
        "nav": f(row.get("nav")),
        "fam_aek": f(row.get("fam_aek")),
        "fam_size": f(row.get("fam_size")),
        "yield_2y": f(row.get("yield_2y")),
        "yield_3y": f(row.get("yield_3y")),
        "sharp_2y": f(row.get("sharp_2y")),
        "sharp_3y": f(row.get("sharp_3y")),
        "ir_2y": f(row.get("ir_2y")),
        "te_2y": f(row.get("te_2y")),
        "alpha_3y": f(row.get("alpha_3y")),
        "beta_3y": f(row.get("beta_3y")),
        "amc_sector_aum_y": f(row.get("amc_sector_aum_y")),
        "amc_sector_score_y": f(row.get("amc_sector_score_y")),
        "fee_amc": f(row.get("fee_amc")),
        "fee_distb": f(row.get("fee_distb")),
        "fee_total": f(row.get("fee_total")),
    })

# 정렬 기준
sector_order = ["국내주식", "국내채권", "해외주식", "해외채권", "혼합형", "기타"]

snapshot = {
    "generated_at": datetime.datetime.now().isoformat(),
    "source_file": SRC.name,
    "filter_defaults": {
        "sharp_2y_min": 1.0,
        "fam_aek_min": 300 * 1e8,
        "require_online": True,
        "require_pension": True,
        "require_normal": True,
    },
    "score_defaults": {
        "aum_weight": 0.25,
        "yield_2y_weight": 0.25,
        "sharp_2y_weight": 0.25,
        "amc_sector_y_weight": 0.25,
    },
    "sector_groups": sector_order,
    "funds": funds,
}

with open(OUT, "w", encoding="utf-8") as f:
    json.dump(snapshot, f, ensure_ascii=False, indent=2)

# 통계
hanto = [f for f in funds if f["amc_nm"] == "한국투자신탁운용"]
print(f"✓ {OUT}")
print(f"  원본 row: {total_rows}, skipped (정상/온라인/퇴직연금 미통과): {skipped}")
print(f"  snapshot 포함: {len(funds)}  (한투 {len(hanto)})")
print(f"  sector 분포: {sorted(sector_set)}")
sec_dist = {sg: sum(1 for f in funds if f["sector_group"] == sg) for sg in sector_order}
print(f"  sector 분포: {sec_dist}")
